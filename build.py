#!/usr/bin/env python3
"""
build.py - Budget Increase Recommendations dashboard builder
Usage: python build.py data.xlsx
Output: index.html (self-contained, commit this to GitHub)
"""
import sys, json, os
from datetime import datetime, date

KEEP_SECTIONS = {
    "Reviewed - No Upgrade Yet",
    "Reviewed - Upgraded 2026",
    "No Longer Counts 2026",
}


def serialize(v):
    if isinstance(v, (datetime, date)):
        return str(v)
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v


def load_xlsx(path):
    try:
        import openpyxl
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
        import openpyxl
    wb = openpyxl.load_workbook(path)
    if 'Raw Data' not in wb.sheetnames:
        raise ValueError(f"Sheet 'Raw Data' not found. Sheets: {wb.sheetnames}")
    ws = wb['Raw Data']
    headers = [cell.value for cell in ws[1]]

    section_idx = None
    for i, h in enumerate(headers):
        if h and str(h).strip() == 'Section':
            section_idx = i
            break

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if section_idx is not None:
            if row[section_idx] not in KEEP_SECTIONS:
                continue
        obj = {h: serialize(row[i]) for i, h in enumerate(headers)
               if h and row[i] is not None and row[i] != ''}
        if obj:
            rows.append(obj)
    return rows


def build(xlsx_path):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, 'template.html')
    output_path = os.path.join(script_dir, 'index.html')
    if not os.path.exists(template_path):
        raise FileNotFoundError("template.html not found in the same folder as build.py")
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"{xlsx_path} not found")

    print(f"Loading {xlsx_path} ...")
    rows = load_xlsx(xlsx_path)
    print(f"  {len(rows)} rows loaded (filtered to the three tracked sections)")

    with open(template_path, 'r', encoding='utf-8') as f:
        html = f.read()

    ts = datetime.now().strftime('%b %d, %Y %I:%M %p')
    json_data = json.dumps(rows, separators=(',', ':'), ensure_ascii=False)
    injection = f"// Embedded dataset - generated {ts}\nconst __EMBEDDED_DATA__={json_data};"
    html = html.replace('// @@DATA_INJECTION@@', injection)
    html = html.replace("DATA_TS='@@TIMESTAMP@@'", f"DATA_TS='{ts}'")

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  index.html written ({os.path.getsize(output_path)//1024} KB)")
    print("Done.")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python build.py data.xlsx")
        sys.exit(1)
    build(sys.argv[1])
